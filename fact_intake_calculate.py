import tkinter.messagebox
from openpyxl.styles import Font
import openpyxl
import datetime
from constants import *
from exceptions import *


def get_fact_intake_calculate(mode, path_to_target_file, progress_var):
    current_year = int(datetime.today().strftime("%Y"))
    current_month = int(datetime.today().strftime("%m"))
    current_day = int(datetime.today().strftime("%d"))

    target_file_formula = openpyxl.load_workbook(path_to_target_file, data_only=False)
    progress_var.put(30)
    try:
        target_sheet_formula = target_file_formula[NAMES_SHEETS_IN_SIMPLE_MOD.get(mode)]
    except KeyError:
        target_file_formula.close()
        raise UnableToFindMainSheetInTargetFile(NAMES_SHEETS_IN_SIMPLE_MOD.get(mode))

    target_file = openpyxl.load_workbook(path_to_target_file, data_only=True)
    progress_var.put(30)
    try:
        rmpa_sheet = target_file["RMPA"]
        target_sheet = target_file[NAMES_SHEETS_IN_SIMPLE_MOD.get(mode)]
    except KeyError:
        target_file.close()
        target_file_formula.close()
        raise UnableToFindMainSheetInTargetFile(f'"{NAMES_SHEETS_IN_SIMPLE_MOD.get(mode)}" или "RMPA"')
    #####################################################################################################
    val = -1

    for i in range(1, target_sheet.max_column):  # Определяем год
        val = target_sheet.cell(row=2, column=i).value
        if val is not None:
            try:
                if int(val) == int(current_year):
                    break
            except ValueError:
                pass
    if not val or int(val) != int(current_year):
        target_file.close()
        target_file_formula.close()
        raise TodayNotFindInSourceFileException

    progress_var.put(6)
    for j in range(i, target_sheet.max_column):  # Определяем месяц
        val = str(target_sheet.cell(row=1, column=j).value).strip()
        if val in ("January", "Январь"):
            val = 1
        elif val in ("February", "Февраль"):
            val = 2
        elif val in ("March", "Март"):
            val = 3
        elif val in ("April", "Апрель"):
            val = 4
        elif val in ("May", "Май"):
            val = 5
        elif val in ("June", "Июнь"):
            val = 6
        elif val in ("July", "Июль"):
            val = 7
        elif val in ("August", "Август"):
            val = 8
        elif val in ("September", "Сентябрь"):
            val = 9
        elif val in ("October", "Октябрь"):
            val = 10
        elif val in ("November", "Ноябрь"):
            val = 11
        elif val in ("December", "Декабрь"):
            val = 12
        if val is not None:
            if val == current_month:
                break

    if not val or val != current_month:
        target_file.close()
        target_file_formula.close()
        raise TodayNotFindInSourceFileException

    progress_var.put(6)
    for k in range(j, j + 27):  # Определяем день
        val = target_sheet.cell(row=3, column=k).value
        if int(val) == int(current_day):
            day = k
            break

    if not val or int(val) != int(current_day):
        target_file.close()
        target_file_formula.close()
        raise TodayNotFindInSourceFileException
    #####################################################################################################
    collected_data = [[], []]

    progress_var.put(6)
    for i in range(2, rmpa_sheet.max_row):

        val_SKU = int(rmpa_sheet.cell(row=i, column=9).value) + int(rmpa_sheet.cell(row=i, column=11).value)
        SKU_self = rmpa_sheet.cell(row=i, column=5).value

        if val_SKU is not None:
            collected_data[0].append(SKU_self)
            collected_data[1].append(val_SKU)
    #####################################################################################################
    progress_var.put(6)
    for i in range(1, target_sheet_formula.max_row):

        val = str(target_sheet.cell(row=i, column=2).value)

        if val in collected_data[0]:
            for j in range(len(collected_data[0])):

                act_sku = collected_data[0][j]
                act_qnt = int(collected_data[1][j])

                if NAMES_SHEETS_IN_SIMPLE_MOD.get(mode) in ("КАРТОН", "ПЛЕНКА"):

                    if str(val) == str(act_sku):
                        if target_sheet.cell(row=i + 2, column=day - 1).value is not None:
                            try:
                                target_sheet_formula.cell(row=i - 1, column=day).value = int(
                                    target_sheet.cell(row=i, column=day - 1).value) + int(
                                    target_sheet.cell(row=i + 2, column=day - 1).value) - act_qnt
                            except TypeError:
                                target_file.close()
                                target_file_formula.close()
                                raise OpenCloseFileException(path_to_target_file)
                        else:
                            if target_sheet.cell(row=i, column=day - 1).value:
                                target_sheet_formula.cell(row=i - 1, column=day).value = int(
                                    target_sheet.cell(row=i, column=day - 1).value) - act_qnt
                        if target_sheet_formula.cell(row=i - 1, column=day).value is not None:
                            if str(target_sheet_formula.cell(row=i - 1, column=day).value)[0] != "=":
                                if int(target_sheet_formula.cell(row=i - 1, column=day).value) < 0:
                                    target_sheet_formula.cell(row=i - 1, column=day).font = Font(
                                        color='7030A0', bold=False, size=9, name="Arial")
                                else:
                                    target_sheet_formula.cell(row=i - 1, column=day).font = Font(color='C00000', bold=False,
                                                                                                 size=9,
                                                                                                 name="Arial")
                        target_sheet_formula.cell(row=i, column=day).font = Font(bold=True, size=9, name="Arial")

                elif NAMES_SHEETS_IN_SIMPLE_MOD.get(mode) == "Сырье":

                    if str(val) == str(act_sku):
                        if target_sheet_formula.cell(row=i + 1, column=day).value is not None:
                            if target_sheet.cell(row=i + 3, column=day - 1).value is not None:
                                target_sheet_formula.cell(row=i + 1, column=day).value = int(
                                    target_sheet.cell(row=i + 2, column=day - 1).value) + target_sheet.cell(row=i + 3,
                                                                                                            column=day - 1).value - act_qnt
                            else:
                                try:
                                    target_sheet_formula.cell(row=i + 1, column=day).value = int(
                                        target_sheet.cell(row=i + 2, column=day - 1).value) - act_qnt
                                except TypeError:
                                    target_file.close()
                                    target_file_formula.close()
                                    raise OpenCloseFileException(path_to_target_file)
                            if target_sheet_formula.cell(row=i + 1, column=day).value is not None:
                                if str(target_sheet_formula.cell(row=i + 1, column=day).value)[0] != "=":
                                    if int(target_sheet_formula.cell(row=i + 1, column=day).value) < 0:
                                        target_sheet_formula.cell(row=i + 1, column=day).font = Font(
                                            color='7030A0', bold=False, size=11, name="Arial")
                                    else:
                                        target_sheet_formula.cell(row=i + 1, column=day).font = Font(color='C00000',
                                                                                                     bold=False, size=11,
                                                                                                     name="Arial")
                        target_sheet_formula.cell(row=i + 2, column=day).font = Font(bold=True, size=11, name="Arial")

    progress_var.put(3)
    target_file_formula.save(path_to_target_file)
