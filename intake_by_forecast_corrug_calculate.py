from openpyxl.styles import Font
import openpyxl
import datetime


def get_contact(day_array, corr, demand, current_year):
    now = datetime.datetime.today().strftime("%d.%m")

    demand = openpyxl.load_workbook(demand, data_only=True)  # Opening source data with order, time, etc. info
    demand_s = demand["Production"]

    corrug = openpyxl.load_workbook(corr)  # Opening target file
    corrug_BOMmini = corrug["BOMmini"]
    corrug_corrug = corrug["КАРТОН"]
    #####################################################################################################
    collected_data = [[], [], []]  # 1

    for i in range(len(day_array[1])):
        val = day_array[0][i]
        col = day_array[1][i]

        for j in range(3, demand_s.max_row):
            val_SKU = demand_s.cell(row=j, column=col).value
            SKU_self = demand_s.cell(row=j, column=1).value

            if val_SKU is not None:
                if val_SKU > 0 and SKU_self not in (None, 0, "0", "Grand Total", "check"):
                    collected_data[0].append(val)
                    collected_data[1].append(SKU_self)
                    collected_data[2].append(val_SKU)
    #####################################################################################################
    target_data = [[], [], []]  # 2

    for i in range(1, corrug_BOMmini.max_row):

        val = corrug_BOMmini.cell(row=i, column=2).value
        req_val = corrug_BOMmini.cell(row=i, column=1).value

        for j in range(len(collected_data[0])):

            temp_val = collected_data[1][j]

            if val == temp_val:
                target_data[0].append(collected_data[0][j])
                target_data[1].append(req_val)
                target_data[2].append(collected_data[2][j])
    #####################################################################################################
    final_data = [[], [], []]  # 3

    for i in range(len(target_data[0])):

        act_data = target_data[0][i]
        act_sku = target_data[1][i]
        act_qnt = 0

        if i > 0:
            for k in range(len(final_data[0])):
                if act_data == final_data[0][k] and act_sku == final_data[1][k]:
                    act_data = '0'
                    act_sku = 0
                    break

        for j in range(len(target_data[0])):

            comparing_data = target_data[0][j]
            comparing_sku = target_data[1][j]

            if act_data == comparing_data and act_sku == comparing_sku:
                act_qnt += target_data[2][j]

        if act_qnt > 0:
            final_data[0].append(target_data[0][i])
            final_data[1].append(target_data[1][i])
            final_data[2].append(act_qnt)
    #####################################################################################################
    req_date = [[], []]  # 4

    for i in range(7, corrug_corrug.max_column):
        val = corrug_corrug.cell(row=2, column=i).value
        if val is not None:
            if int(val) == current_year:
                break

    for j in range(i, corrug_corrug.max_column):
        val = corrug_corrug.cell(row=2, column=j).value
        if val in final_data[0] and val not in req_date[1]:
            req_date[0].append(j)
            req_date[1].append(val)

    # for p in range(len(final_data[0])):
    # print(final_data[0][p], final_data[1][p], final_data[2][p])
    #####################################################################################################
    for i in range(1, corrug_corrug.max_row):  # 5
        print(i)

        val = corrug_corrug.cell(row=i, column=2).value

        if str(val) in str(final_data[1]):
            # print(i)

            for j in range(len(final_data[1])):

                act_sku = final_data[1][j]
                act_date = final_data[0][j]
                act_qnt = final_data[2][j]

                if str(val) == str(act_sku):
                    for d in range(7):
                        if corrug_corrug.cell(row=i - 3,
                                              column=req_date[0][
                                                         req_date[1].index(
                                                             act_date)] + d).value is None and "TOTAL" not in str(
                            corrug_corrug.cell(row=1, column=req_date[0][req_date[1].index(
                                act_date)] + d).value):
                            corrug_corrug.cell(row=i - 3,
                                               column=req_date[0][
                                                          req_date[1].index(act_date)] + d).value = str(
                                act_qnt) + " |" + str(now)
                            break

                    corrug_corrug.cell(row=i - 3,
                                       column=req_date[0][req_date[1].index(act_date)] + d).font = Font(
                        color='00008B', bold=False, size=9, name="Arial")

                    # Тут начинается точечный расход по дням
                    additional_row = False  # булеан для разделителя месяцев
                    for h in range(3, 7):
                        if "TOTAL" in str(corrug_corrug.cell(row=1, column=req_date[0][req_date[1].index(
                                act_date)] + h).value):
                            additional_row = True
                            continue
                        corrug_corrug.cell(row=i - 1,
                                           column=req_date[0][req_date[1].index(act_date)] + h).value = int(
                            act_qnt) / 4
                        corrug_corrug.cell(row=i - 1,
                                           column=req_date[0][req_date[1].index(act_date)] + h).font = Font(
                            color='00008B', bold=False, size=9, name="Arial")
                    if additional_row:
                        corrug_corrug.cell(row=i - 1,
                                           column=req_date[0][
                                                      req_date[1].index(act_date)] + h + 1).value = int(
                            act_qnt) / 4
                        corrug_corrug.cell(row=i - 1,
                                           column=req_date[0][
                                                      req_date[1].index(act_date)] + h + 1).font = Font(
                            color='00008B', bold=False, size=9, name="Arial")

    corrug.save(corr)
