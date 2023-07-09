from openpyxl.styles import Font
import openpyxl
import datetime


def get_postfactum(corr):
    
    current_year = int(datetime.datetime.today().strftime("%Y"))
    current_month = int(datetime.datetime.today().strftime("%m"))
    current_day = int(datetime.datetime.today().strftime("%d"))

    corrug_formula = openpyxl.load_workbook(corr, data_only=False)
    try:
        corrug_f = corrug_formula["КАРТОН"]
    except KeyError:
        corrug_f = corrug_formula["Сырье"]

    corrug = openpyxl.load_workbook(corr, data_only=True)  # Opening target file
    corrug_RMPA = corrug["RMPA"]
    try:
        corrug_corrug = corrug["КАРТОН"]
    except KeyError:
        corrug_corrug = corrug["Сырье"]
#####################################################################################################
    for i in range(7, corrug_corrug.max_column):  # Определяем год
        val = corrug_corrug.cell(row=2, column=i).value
        if val is not None:
            if int(val) == int(current_year):
                break

    for j in range(i, corrug_corrug.max_column):  # Определяем месяц
        val = corrug_corrug.cell(row=1, column=j).value
        if val in ("January", "Январь"):
            val = 1
        elif val in ("February", "Февраль"):
            val = 2
        elif val in ("March", "Март"):
            val = 3
        elif val in ("April", "Апрель"):
            val = 4
        elif val in ("May", "Май"):
            val = 5
        elif val in ("June", "Июнь"):
            val = 6
        elif val in ("July", "Июль"):
            val = 7
        elif val in ("August", "Август"):
            val = 8
        elif val in ("September", "Сентябрь"):
            val = 9
        elif val in ("October", "Октябрь"):
            val = 10
        elif val in ("November", "Ноябрь"):
            val = 11
        elif val in ("December", "Декабрь"):
            val = 12
        if val is not None:
            if val == current_month:
                break

    for k in range(j, j + 27):  # Определяем день
        val = corrug_corrug.cell(row=3, column=k).value
        if int(val) == int(current_day):
            day = k
            break
#####################################################################################################
    collected_data = [[], []]

    for i in range(2, corrug_RMPA.max_row):

        val_SKU = int(corrug_RMPA.cell(row=i, column=9).value) + int(corrug_RMPA.cell(row=i, column=11).value)
        SKU_self = corrug_RMPA.cell(row=i, column=5).value

        if val_SKU is not None:
            collected_data[0].append(SKU_self)
            collected_data[1].append(val_SKU)
#####################################################################################################
    for i in range(1, corrug_f.max_row):

        val = str(corrug_corrug.cell(row=i, column=2).value)

        if val in collected_data[0]:
            for j in range(len(collected_data[0])):

                act_sku = collected_data[0][j]
                act_qnt = int(collected_data[1][j])
                
                # For corrug

                if 'КАРТОН' in corrug.get_sheet_names():

                    if str(val) == str(act_sku):
                        if corrug_corrug.cell(row=i + 2, column=day - 1).value is not None:
                            corrug_f.cell(row=i - 1, column=day).value = int(corrug_corrug.cell(row=i, column=day - 1).value) + int(corrug_corrug.cell(row=i + 2, column=day - 1).value) - act_qnt
                        else:
                            corrug_f.cell(row=i - 1, column=day).value = int(corrug_corrug.cell(row=i, column=day - 1).value) - act_qnt
                        if corrug_f.cell(row=i - 1, column=day).value is not None:
                            if int(corrug_f.cell(row=i - 1, column=day).value) < 0:
                                corrug_f.cell(row=i - 1, column=day).font = Font(
                                    color='7030A0', bold=False, size=9, name="Arial")
                            else:
                                corrug_f.cell(row=i - 1, column=day).font = Font(color='C00000', bold=False, size=9, name="Arial")
                        corrug_f.cell(row=i, column=day).font = Font(bold=True, size=9, name="Arial")

                    # For raw

                elif str(val) == str(act_sku):
                    if corrug_f.cell(row=i + 1, column=day).value is not None:
                        if corrug_corrug.cell(row=i + 3, column=day - 1).value is not None:
                            corrug_f.cell(row=i + 1, column=day).value = int(
                                corrug_corrug.cell(row=i + 2, column=day - 1).value) + corrug_corrug.cell(row=i + 3, column=day - 1).value - act_qnt
                        else:
                            corrug_f.cell(row=i + 1, column=day).value = int(
                                corrug_corrug.cell(row=i + 2, column=day - 1).value) - act_qnt
                        if corrug_f.cell(row=i + 1, column=day).value is not None:
                            if int(corrug_f.cell(row=i + 1, column=day).value) < 0:
                                corrug_f.cell(row=i + 1, column=day).font = Font(
                                    color='7030A0', bold=False, size=11, name="Arial")
                            else:
                                corrug_f.cell(row=i + 1, column=day).font = Font(color='C00000', bold=False, size=11,
                                                                                 name="Arial")
                    corrug_f.cell(row=i + 2, column=day).font = Font(bold=True, size=11, name="Arial")

    corrug_formula.save(corr)
