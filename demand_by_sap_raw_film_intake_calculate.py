import os
from openpyxl.styles import Font
import openpyxl
import datetime
from constants import *
from exceptions import *


def get_demand_by_sap_raw_film_intake_calculate(mode, day_array, path_to_target_file, path_to_source_dir, progress_var):
    main_offset = 2 if mode == 5 else 4
    detail_offset = 1 if mode == 5 else -1
    font_size = 11 if mode == 5 else 8
    now = datetime.today().strftime("%d.%m")
    target_file = openpyxl.load_workbook(path_to_target_file, data_only=False)
    progress_var.put(34)
    try:
        target_file_sheet = target_file[SAP_DEMAND_MODS.get(mode)]
    except KeyError:
        target_file.close()
        raise UnableToFindMainSheetInTargetFile(SAP_DEMAND_MODS.get(mode))
    required_dates = [[], []]

    progress_var.put(3)
    for i in range(3, target_file_sheet.max_column):  # detecting coordinates in target_file file (year)
        val = target_file_sheet.cell(row=2, column=i).value
        if val:
            try:
                if int(val) in day_array[0]:
                    break
            except:
                pass

    progress_var.put(3)
    for j in range(i, target_file_sheet.max_column):  # detecting coordinates in target_file file (week)
        try:
            val = int(target_file_sheet.cell(row=2, column=j).value)
        except:
            continue
        if val in day_array[1] and val not in required_dates[1]:
            required_dates[0].append(j)
            required_dates[1].append(str(val))

    progress_var.put(3)
    for q in range(len(day_array[0])):
        sap_demand_file_name = path_to_source_dir + r"\\"[0] + str(day_array[0][q]) + "-" + str(day_array[1][q]) + ".XLSX"
        sap_demand_file = openpyxl.load_workbook(sap_demand_file_name, data_only=True)
        progress_var.put(2)
        sap_demand_file_sheet = sap_demand_file.worksheets[0]
        collected_data = [[], [], []]

        for i in range(2, sap_demand_file_sheet.max_row):  # collecting SKU-consumption from SAP load
            curr_sku = sap_demand_file_sheet.cell(row=i, column=1).value
            curr_quant = int(sap_demand_file_sheet.cell(row=i, column=4).value)
            if curr_quant >= 0:
                collected_data[0].append(str(day_array[1][q]))
                collected_data[1].append(curr_sku)
                collected_data[2].append(curr_quant)

        for i in range(2, target_file_sheet.max_row):  # 5

            val = target_file_sheet.cell(row=i, column=2).value

            if str(val) in str(collected_data[1]):
                # print(i)

                for j in range(len(collected_data[1])):

                    act_sku = collected_data[1][j]
                    act_date = collected_data[0][j]
                    act_qnt = collected_data[2][j]

                    if str(val) == str(act_sku):
                        for d in range(3):
                            if target_file_sheet.cell(row=i - main_offset,
                                                      column=required_dates[0][
                                                                 required_dates[1].index(
                                                                     act_date)] + d).value is None and "TOTAL" not in str(
                                target_file_sheet.cell(row=1, column=required_dates[0][required_dates[1].index(
                                    act_date)] + d).value):
                                target_file_sheet.cell(row=i - main_offset,
                                                       column=required_dates[0][
                                                                  required_dates[1].index(act_date)] + d).value = str(
                                    act_qnt) + " |" + str(now)
                                break

                        target_file_sheet.cell(row=i - main_offset,
                                               column=required_dates[0][
                                                          required_dates[1].index(act_date)] + d).font = Font(
                            color='00008B', bold=False, size=font_size, name="Arial")

                        # Тут начинается точечный расход по дням
                        additional_row = False  # булеан для разделителя месяцев
                        if act_qnt > 0:
                            for h in range(3, 7):
                                if "TOTAL" in str(
                                        target_file_sheet.cell(row=1, column=required_dates[0][required_dates[1].index(
                                            act_date)] + h).value):
                                    additional_row = True
                                    continue
                                target_file_sheet.cell(row=i + detail_offset,
                                                       column=required_dates[0][
                                                                  required_dates[1].index(act_date)] + h).value = int(
                                    act_qnt) / 4
                                target_file_sheet.cell(row=i + detail_offset,
                                                       column=required_dates[0][
                                                                  required_dates[1].index(act_date)] + h).font = Font(
                                    color='00008B', bold=False, size=font_size, name="Arial")
                            if additional_row:
                                target_file_sheet.cell(row=i + detail_offset,
                                                       column=required_dates[0][
                                                                  required_dates[1].index(
                                                                      act_date)] + h + 1).value = int(
                                    act_qnt) / 4
                                target_file_sheet.cell(row=i + detail_offset,
                                                       column=required_dates[0][
                                                                  required_dates[1].index(
                                                                      act_date)] + h + 1).font = Font(
                                    color='00008B', bold=False, size=font_size, name="Arial")

    target_file.save(path_to_target_file)
