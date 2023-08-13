import os
import re
import shutil
import time
import traceback
import zipfile
from pathlib import Path
from tkinter import messagebox
import openpyxl
from constants import *
from exceptions import *
from intake_by_plan_corrug_calculate import get_intake_by_plan_corrug_calculate
from intake_by_forecast_corrug_calculate import get_intake_by_forecast_corrug_calculate
from fact_intake_calculate import get_fact_intake_calculate
from demand_by_sap_raw_film_intake_calculate import get_demand_by_sap_raw_film_intake_calculate


def get_paths_to_files():
    # get paths to main files

    paths_to_files = {
        "картон.xl": "",
        "пленка.xl": "",
        "сырье.xl": "",
        "production": "",
        "план": "",
        "weeklyintakebysap": ""
    }
    Path("./" + NAME_OF_DIR_SAP_DEMAND).mkdir(parents=True, exist_ok=True)
    for pattern in paths_to_files.keys():
        for file in os.listdir():
            if re.match(pattern, file.lower()):
                paths_to_files[pattern] = os.getcwd() + r"\\"[0] + file

    return paths_to_files


def get_corrug_calculating_days_from_plan(path_to_file):
    plan_wb = openpyxl.load_workbook(filename=path_to_file, data_only=True)
    try:
        plan_wb_s = plan_wb["Plan"]
    except KeyError:
        raise UnableToFindExcelSheet("Plan")
    day_array = [[], [], [], [], []]
    today_value = ""

    for i in range(1, plan_wb_s.max_column):
        today_value = plan_wb_s.cell(row=3, column=i).value
        if today_value and str(today_value) == str(TODAY):
            break
    if str(today_value) != str(TODAY):
        plan_wb.close()
        raise TodayNotFindInSourceFileException("Не удалось найти ячейку текущего дня в файле плана.")

    for j in range(i, plan_wb_s.max_column):
        val = plan_wb_s.cell(row=3, column=j).value
        if val is not None:
            day_array[0].append(int(val.strftime("%Y")))
            day_array[1].append(int(val.strftime("%m")))
            day_array[2].append(int(val.strftime("%d")))
            day_array[3].append(int(j))
            day_array[4].append("")


    plan_wb.close()
    return day_array


def get_corrug_calculating_weeks_from_forecast(path_to_file):
    forecast_wb = openpyxl.load_workbook(filename=path_to_file, data_only=True)
    try:
        forecast_wb_s = forecast_wb["Production"]
    except KeyError:
        raise UnableToFindExcelSheet("Production")
    day_array = [[], [], [], [], []]
    val = ""

    for i in range(1, forecast_wb_s.max_column):
        val = forecast_wb_s.cell(row=2, column=i).value
        if val is not None and str(val) == str(CURRENT_WEEK):
            break
    if str(val) != str(CURRENT_WEEK):
        forecast_wb.close()
        raise TodayNotFindInSourceFileException("Не удалось найти ячейку текущей недели в файле прогноза. "
                                                "Возможно файл старый")

    for j in range(i, forecast_wb_s.max_column):
        val = forecast_wb_s.cell(row=2, column=j).value
        if val is not None:
            day_array[0].append(val)
            day_array[1].append(j)
            day_array[2].append(CURRENT_YEAR)
            day_array[3].append(0)
            day_array[4].append("")

    forecast_wb.close()
    return day_array


def get_film_raw_week_year_from_sap_demand(path_to_file):
    day_array = [[], [], [], [], []]
    for file in os.listdir(path=path_to_file):
        if re.match(r"20\d{2}-\d{2}.(xl|XL)", file):
            day_array[0].append(str(file[0:4]))
            day_array[1].append(str(file[5:7]))
            day_array[2].append(0)
            day_array[3].append(0)
            day_array[4].append("")

    return day_array


def calculate_and_echo_to_target_file(mode, selected_days_arr, path_to_target_file,
                                      path_to_source_file, selected_year, progress_var):
    # reserve copy first

    Path("./ReserveCopy").mkdir(parents=True, exist_ok=True)
    progress_var.put(5)
    shutil.copyfile(path_to_target_file, "./ReserveCopy/" + "reserve_copy_" + os.path.basename(path_to_target_file))
    if mode not in SIMPLE_MODS and mode not in SAP_DEMAND_MODS.keys():
        shutil.copyfile(path_to_source_file, "./ReserveCopy/" + "reserve_copy_" + os.path.basename(path_to_source_file))
    progress_var.put(5)

    try:
        if mode == 0:
            get_intake_by_plan_corrug_calculate(selected_days_arr[0:-1], path_to_target_file, path_to_source_file,
                                                selected_year, progress_var)

        elif mode == 1:
            get_intake_by_forecast_corrug_calculate((selected_days_arr[-1], selected_days_arr[-2]),
                                                    path_to_target_file, path_to_source_file,
                                                    selected_year, progress_var)

        elif mode in SAP_DEMAND_MODS.keys():
            get_demand_by_sap_raw_film_intake_calculate(mode, (selected_days_arr[0], selected_days_arr[4]),
                                                        path_to_target_file, path_to_source_file, progress_var)

        elif mode in SIMPLE_MODS:
            get_fact_intake_calculate(mode, path_to_target_file, progress_var)

        progress_var.put(100)

    except UnableToFindMainSheetInTargetFile as e:
        e_str = str(e)
        progress_var.put(
            lambda: messagebox.showwarning("Внимание!", f'Не удалось найти вкладку "{e_str}" в целевом файле. '
                                                        'Возможно она была переименована.'))

    except UnableToFindBomSheetInTargetFile as e:
        e_str = str(e)
        progress_var.put(
            lambda: messagebox.showwarning("Внимание!", f'Не удалось найти вкладку "{e_str}" в целевом файле. '
                                                        'Возможно она была переименована.'))

    except (PermissionError, zipfile.BadZipfile):
        progress_var.put(lambda: messagebox.showwarning("Внимание!",
                                                        f'Не удалось открыть или закрыть целевой или файл-ресурс. '
                                                        f'Закройте файлы,'
                                                        f' если они открыты, или проверьте их целостность.'))

    except TodayNotFindInSourceFileException:
        progress_var.put(lambda: messagebox.showwarning("Внимание!",
                                                        f'Не удалось найти текущий день (неделю, месяц, год) в целевом '
                                                        f'файле. Убедитесь, что количество дней (месяцев, недель, лет) '
                                                        f'в файле-источнике такое же как и в целевом файле.'))

    except OpenCloseFileException as e:
        e_str = str(e)
        progress_var.put(lambda: messagebox.showwarning("Внимание!",
                                                        f'Возникла техническая ошибка. Чтобы её исправить откройте '
                                                        f'файл "{e_str}", просто сохраните и закройте его.'))

    except Exception:
        e_str = traceback.format_exc()
        progress_var.put(lambda: messagebox.showerror("Критическая ошибка!", e_str))
        exit(1)
