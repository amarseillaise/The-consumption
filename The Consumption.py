import tkinter as tk
from tkinter import filedialog, Checkbutton, BooleanVar, messagebox
from tkinter.ttk import Label, Combobox
import os
import datetime
from intake_by_plan_corrug_calculate import get_intake_by_plan_corrug_calculate
from intake_by_forecast_corrug_calculate import get_intake_by_forecast_corrug_calculate
from fact_intake_calculate import get_fact_intake_calculate
from demand_by_sap_raw_film_intake_calculate import get_demand_by_sap_raw_film_intake_calculate

import openpyxl

w = 33
par_dir = os.path.abspath(os.path.join(os.getcwd(), os.pardir))
today = datetime.datetime.today().strftime("%Y-%m-%d") + " 00:00:00"
current_week = int(datetime.datetime.today().strftime("%W"))
checks = []
day_array = [[], [], [], []]
req_days = [[], [], [], [], []]
weeks_for_raw_demand = [[], []]

for corrug_path in os.listdir(path='..'):
    if 'КАРТОН' in corrug_path:
        corrug_sheet = "КАРТОН"
        corrug_path = par_dir + "/" + corrug_path
        break
    corrug_path = None
    corrug_sheet = None
for raw_path in os.listdir(path='..'):
    if 'СЫРЬЕ' in raw_path:
        raw_sheet = "Сырье"
        raw_path = par_dir + "/" + raw_path
        break
    raw_path = None
    raw_sheet = None
for plan_path in os.listdir(path='..'):
    if 'План' in plan_path:
        plan_sheet = "Plan"
        plan_path = par_dir + "/" + plan_path
        break
    plan_path = None
    plan_sheet = None
for demand_path in os.listdir(path='..'):
    if 'Production' in demand_path:
        demand_sheet = "Production"
        demand_path = par_dir + "/" + demand_path
        break
    demand_path = None
    demand_sheet = None

for j in os.listdir(path='WeeklyIntakeBySAP'):  # collecting week and day from folder with SAP_load
    weeks_for_raw_demand[0].append(str(j[0:4]))
    weeks_for_raw_demand[1].append(str(j[5:7]))


class CheckButton:
    def __init__(self, space, title, master):
        self.var = BooleanVar()
        self.var.set(1)
        self.title = title
        self.cb = Checkbutton(
            master,
            text=title,
            variable=self.var,
            onvalue=1,
            offvalue=0
        )
        self.cb.place(
            x=115 + float(space) * 35,
            rely=0.15,
            height=14
        )


class Button:
    def __init__(self, title, master, func):
        self.title = title
        self.b = tk.Button(
            master,
            text=title,
            width=10,
            command=func
        )
        self.b.place(
            relx=0.82,
            rely=0.1,
        )


class msgbox:
    def __init__(self, title, message):
        self.title = title
        self.message = message
        self.msg = messagebox.showinfo(
            title=title,
            message=message
        )


class ermsgbox:
    def __init__(self, title, message):
        self.title = title
        self.message = message
        self.msg = messagebox.showerror(
            title=title,
            message=message
        )


def ident_days(window_sel, xlsx, xlsx_s):
    global day_array, checks
    if len(checks) > 0:
        checks = []
        ch_destroy()
    plan_wb = openpyxl.load_workbook(filename=xlsx, data_only=True)
    plan_wb_s = plan_wb[xlsx_s]
    day_array = [[], [], [], []]

    for i in range(1, plan_wb_s.max_column):
        val = plan_wb_s.cell(row=3, column=i).value
        if val is not None and str(val) == str(today):
            break

    for j in range(i, plan_wb_s.max_column):
        val = plan_wb_s.cell(row=3, column=j).value
        if val is not None:
            day_array[0].append(int(val.strftime("%Y")))
            day_array[1].append(int(val.strftime("%m")))
            day_array[2].append(int(val.strftime("%d")))
            day_array[3].append(int(j))

    for p in range(len(day_array[0])):
        checks.append(CheckButton(p, day_array[2][p], window_sel))
    return day_array


def ch_destroy():
    for ch in checks:
        ch.cb.destroy()


def transfer_data(choise, list, arr, file_src, file_trg, destroy):
    global req_days
    i = 0
    for h in list:
        if h.var.get():
            req_days[0].append(arr[0][i])
            req_days[1].append(arr[1][i])
            req_days[2].append(arr[2][i])
            req_days[3].append(arr[3][i])
        i += 1
    if choise == 1:
        get_intake_by_plan_corrug_calculate(req_days, file_trg, file_src, int(VarYear.get()))
    elif choise == 2:
        get_intake_by_forecast_corrug_calculate(req_days, file_trg, file_src, int(VarYear.get()))
    elif choise == 3:
        try:
            get_fact_intake_calculate(file_trg)
        except TypeError:
            ermsgbox("Ошибка", "Сохрани целевой файл и запусти заново. Или неправильная выгрузка RMPA")
            destroy.destroy()
            raise SystemExit
    elif choise == 4:
        get_demand_by_sap_raw_film_intake_calculate(file_trg)
    elif choise == 5:
        try:
            get_fact_intake_calculate(file_trg)
        except TypeError:
            ermsgbox("Ошибка", "Сохрани целевой файл и запусти заново. Или неправильная выгрузка RMPA")
            destroy.destroy()
            raise SystemExit
    msgbox('Грума гордится тобой', 'Рассчёт окончен')
    destroy.destroy()


def ident_weeks(window_sel, xlsx, xlsx_s):
    global day_array, checks
    if len(checks) > 0:
        ch_destroy()
        checks = []
    plan_wb = openpyxl.load_workbook(filename=xlsx, data_only=True)
    plan_wb_s = plan_wb[xlsx_s]
    day_array = [[], [], [], []]

    for i in range(1, plan_wb_s.max_column):
        val = plan_wb_s.cell(row=2, column=i).value
        if val is not None and str(val) == str(current_week):
            break

    for j in range(i, plan_wb_s.max_column):
        val = plan_wb_s.cell(row=2, column=j).value
        if val is not None:
            day_array[0].append(val)
            day_array[1].append(j)
            day_array[2].append(0)
            day_array[3].append(0)

    for p in range(len(day_array[0])):
        checks.append(CheckButton(p, day_array[0][p], window_sel))
    return day_array,


def plan_window():
    corrug_window.destroy()
    window_pp = tk.Tk()
    window_pp.geometry('900x100')
    window_pp.resizable(False, False)
    window_pp.title('Рассчёт расхода по плану производства')

    # Кнопочки
    def select_corrug():
        global corrug_path
        corrug_path = filedialog.askopenfilename(
            initialdir="..",
            title="Выбрать целевой файл",
            filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xlsm"), ("all files", "*.*"))
        )
        corrug_lbl = Label(
            window_pp,
            text=corrug_path,
            width=100
        )
        corrug_lbl.place(
            relx=0.36,
            rely=0.7,
        )
        return corrug_path

    def select_plan():
        global plan_path

        plan_path = filedialog.askopenfilename(
            initialdir="..",
            title="Выбрать файл План",
            filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xlsm"), ("all files", "*.*"))
        )
        plan = Label(
            window_pp,
            text=plan_path,
            width=100
        )
        plan.place(
            relx=0.36,
            rely=0.4,
        )
        if '.xls' in plan_path:
            ident_days(window_pp, plan_path, "Plan")
        return plan_path

    select_plan_file = tk.Button(
        window_pp,
        text="Выбрать файл План",
        width=20,
        command=select_plan
    )
    select_plan_file.place(
        relx=0.05,
        rely=0.35,
    )

    select_target_file = tk.Button(
        window_pp,
        text="Выбрать целевой файл",
        width=20,
        command=select_corrug
    )
    select_target_file.place(
        relx=0.05,
        rely=0.65,
    )
    # Ярлычки и чекбоксы
    corrug_lbl = Label(
        window_pp,
        text=corrug_path,
        width=100
    )
    corrug_lbl.place(
        relx=0.36,
        rely=0.7,
    )

    plan_lbl = Label(
        window_pp,
        text=plan_path,
        width=100
    )
    plan_lbl.place(
        relx=0.36,
        rely=0.4,
    )
    dof_label = Label(
        window_pp,
        text='Дни месяца:',
        width=13
    )
    dof_label.place(
        relx=0.05,
        rely=0.12,
    )
    Button("Посчитать", window_pp, lambda: transfer_data(1, checks, day_array, plan_path, corrug_path, window_pp))
    if plan_path is not None:
        if '.xls' in plan_path:
            ident_days(window_pp, plan_path, "Plan")
    window_pp.mainloop()


def demand_window():
    corrug_window.destroy()
    window_d = tk.Tk()
    window_d.geometry('500x100')
    window_d.resizable(False, False)
    window_d.title('Рассчёт расхода по прогнозу')

    # Кнопочки
    def select_corrug():
        global corrug_path
        corrug_path = filedialog.askopenfilename(
            initialdir="..",
            title="Выбрать целевой файл",
            filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xlsm"), ("all files", "*.*"))
        )
        corrug_lbl = Label(
            window_d,
            text=corrug_path,
            width=100
        )
        corrug_lbl.place(
            relx=0.36,
            rely=0.7,
        )
        return corrug_path

    def select_demand():
        global demand_path

        demand_path = filedialog.askopenfilename(
            initialdir="..",
            title="Выбрать файл прогноз",
            filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xlsm"), ("all files", "*.*"))
        )
        demand = Label(
            window_d,
            text=demand_path,
            width=100
        )
        demand.place(
            relx=0.36,
            rely=0.4,
        )
        if '.xls' in demand_path:
            ident_weeks(window_d, demand_path, "Production")
        return demand_path

    select_demand_file = tk.Button(
        window_d,
        text="Выбрать файл прогноз",
        width=20,
        command=select_demand
    )
    select_demand_file.place(
        relx=0.05,
        rely=0.35,
    )

    select_target_file = tk.Button(
        window_d,
        text="Выбрать целевой файл",
        width=20,
        command=select_corrug
    )
    select_target_file.place(
        relx=0.05,
        rely=0.65,
    )
    # Ярлычки и чекбоксы
    corrug_lbl = Label(
        window_d,
        text=corrug_path,
        width=100
    )
    corrug_lbl.place(
        relx=0.36,
        rely=0.7,
    )

    demand_lbl = Label(
        window_d,
        text=demand_path,
        width=100
    )
    demand_lbl.place(
        relx=0.36,
        rely=0.4,
    )
    dof_label = Label(
        window_d,
        text='Недели года:',
        width=13
    )
    dof_label.place(
        relx=0.05,
        rely=0.12,
    )
    warning_label = Label(
        window_d,
        text='Убедись, что количестов недель в целевом файле не меньше, чем в прогнозе',
        width=80,
        font="Arial 7",
        foreground="red"
    )
    warning_label.place(
        relx=0.12,
        rely=0.00,
        height=11.5
    )
    Button("Посчитать", window_d, lambda: transfer_data(2, checks, day_array, demand_path, corrug_path, window_d))
    if demand_path is not None:
        if '.xls' in demand_path:
            ident_weeks(window_d, demand_path, "Production")
    window_d.mainloop()


def ps_window():
    corrug_window.destroy()
    window_c = tk.Tk()
    window_c.geometry('500x100')
    window_c.resizable(False, False)
    window_c.title('Рассчёт расхода постфактум')

    # Кнопочки
    def select_corrug():
        global corrug_path
        corrug_path = filedialog.askopenfilename(
            initialdir="..",
            title="Выбрать целевой файл",
            filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xlsm"), ("all files", "*.*"))
        )
        corrug_lbl = Label(
            window_c,
            text=corrug_path,
            width=100
        )
        corrug_lbl.place(
            relx=0.36,
            rely=0.7,
        )
        return corrug_path

    select_target_file = tk.Button(
        window_c,
        text="Выбрать целевой файл",
        width=20,
        command=select_corrug
    )
    select_target_file.place(
        relx=0.05,
        rely=0.65,
    )
    # Ярлычки и чекбоксы
    corrug_lbl = Label(
        window_c,
        text=corrug_path,
        width=100
    )
    corrug_lbl.place(
        relx=0.36,
        rely=0.7,
    )
    Button("Посчитать", window_c, lambda: transfer_data(3, checks, day_array, demand_path, corrug_path, window_c))
    window_c.mainloop()


def corr_window():
    global corrug_window
    main_window.destroy()
    corrug_window = tk.Tk()
    corrug_window.geometry('300x200')
    corrug_window.resizable(False, False)
    corrug_window.title('Рассчёт расхода упаковки')

    prod_plan = tk.Button(
        corrug_window,
        text="Рассчитать расход по плану производства",
        width=w,
        command=plan_window
    )
    prod_plan.pack(
        ipadx=5,
        ipady=5,
        expand=True
    )

    demand = tk.Button(
        corrug_window,
        text="Рассчитать расход по прогнозу",
        width=w,
        command=demand_window
    )
    demand.pack(
        ipadx=5,
        ipady=5,
        expand=True
    )

    consumption = tk.Button(
        corrug_window,
        text="Рассчитать расход постфактум",
        width=w,
        command=ps_window
    )
    consumption.pack(
        ipadx=5,
        ipady=5,
        expand=True
    )
    corrug_window.mainloop()


def raw_window():

    main_window.destroy()
    global raw_window
    raw_window = tk.Tk()
    raw_window.geometry('300x200')
    raw_window.resizable(False, False)
    raw_window.title('Рассчёт расхода сырья')

    demand_plan_btn = tk.Button(
        raw_window,
        text="Рассчитать расход по прогнозу",
        width=w,
        command=raw_demand_window
    )
    demand_plan_btn.pack(
        ipadx=5,
        ipady=5,
        expand=True
    )

    ps_btn = tk.Button(
        raw_window,
        text="Рассчитать расход постфактум",
        width=w,
        command=ps_raw_window
    )
    ps_btn.pack(
        ipadx=5,
        ipady=5,
        expand=True
    )


def raw_demand_window():
    raw_window.destroy()
    raw_demand_window = tk.Tk()
    raw_demand_window.geometry('500x100')
    raw_demand_window.resizable(False, False)
    raw_demand_window.title('Рассчитать расход сырья по прогнозу')

    # Кнопочки
    def select_raw():
        global raw_path
        raw_path = filedialog.askopenfilename(
            initialdir="..",
            title="Выбрать целевой файл",
            filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xlsm"), ("all files", "*.*"))
        )
        raw_lbl = Label(
            raw_demand_window,
            text=raw_path,
            width=100
        )
        raw_lbl.place(
            relx=0.36,
            rely=0.7,
        )
        return raw_path

    select_target_file = tk.Button(
        raw_demand_window,
        text="Выбрать целевой файл",
        width=20,
        command=select_raw
    )
    select_target_file.place(
        relx=0.05,
        rely=0.65,
    )
    # Ярлычки и чекбоксы
    raw_lbl = Label(
        raw_demand_window,
        text=raw_path,
        width=100
    )
    raw_lbl.place(
        relx=0.36,
        rely=0.7,
    )
    raw_demand_lbl = Label(
        raw_demand_window,
        text=f'Подгруженные недели: {weeks_for_raw_demand[1]}'
    )
    raw_demand_lbl.place(
        relx=0.05,
        rely=0.3
    )
    Button("Посчитать", raw_demand_window, lambda: transfer_data(4, checks, day_array, demand_path, raw_path, raw_demand_window))
    raw_demand_window.mainloop()
    

def ps_raw_window():
    raw_window.destroy()
    window_c_raw = tk.Tk()
    window_c_raw.geometry('500x100')
    window_c_raw.resizable(False, False)
    window_c_raw.title('Рассчёт расхода сырья постфактум')

    # Кнопочки
    def select_raw():
        global raw_path
        raw_path = filedialog.askopenfilename(
            initialdir="..",
            title="Выбрать целевой файл",
            filetypes=(("Excel files", "*.xlsx"), ("Excel files", "*.xlsm"), ("all files", "*.*"))
        )
        raw_lbl = Label(
            window_c_raw,
            text=corrug_path,
            width=100
        )
        raw_lbl.place(
            relx=0.36,
            rely=0.7,
        )
        return corrug_path

    select_target_file = tk.Button(
        window_c_raw,
        text="Выбрать целевой файл",
        width=20,
        command=select_raw
    )
    select_target_file.place(
        relx=0.05,
        rely=0.65,
    )
    # Ярлычки и чекбоксы
    raw_lbl = Label(
        window_c_raw,
        text=raw_path,
        width=100
    )
    raw_lbl.place(
        relx=0.36,
        rely=0.7,
    )
    Button("Посчитать", window_c_raw, lambda: transfer_data(5, checks, day_array, raw_path, raw_path, window_c_raw))
    window_c_raw.mainloop()
    
main_window = tk.Tk()
main_window.geometry('300x200')
main_window.resizable(False, False)
main_window.title('Выберите категорию расхода')

corrug_consumption_btn = tk.Button(
    main_window,
    text="Упаковка",
    width=w - 15,
    command=corr_window
)
corrug_consumption_btn.pack(
    ipadx=5,
    ipady=5,
    expand=True
)

raw_consumption_btn = tk.Button(
    main_window,
    text="Сырьё",
    width=w - 15,
    command=raw_window
)
raw_consumption_btn.pack(
    ipadx=5,
    ipady=5,
    expand=True
)
VarYear = tk.StringVar()
year_box = Combobox(
    main_window,
    textvariable=VarYear
)
year_box['values'] = (int(datetime.datetime.today().strftime("%Y")) - 1, int(datetime.datetime.today().strftime("%Y")), int(datetime.datetime.today().strftime("%Y")) + 1)
year_box.current(1)
year_box.pack(
    ipadx=5,
    ipady=5,
    expand=True
)
main_window.mainloop()
