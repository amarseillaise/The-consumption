from openpyxl.styles import Font, Color
import openpyxl
import datetime
from exceptions import  *


def get_intake_by_plan_corrug_calculate(day_array, corr, plan, current_year):  # day_array[[y], [m], [d], [col]]
    req_date = [[], []]
    current_week = int(datetime.datetime.today().strftime("%W"))
    current_year_real = int(datetime.datetime.today().strftime("%Y"))

    corrug = openpyxl.load_workbook(filename=corr)  # Opening target file
    try:
        corrug_BOMmini = corrug["BOMmini"]
    except KeyError as e:
        corrug.close()
        raise UnableToFindBomSheetInTargetFile("BOMmini")

    try:
        corrug_corrug = corrug["КАРТОН"]
    except KeyError as e:
        corrug.close()
        raise UnableToFindMainSheetInTargetFile("КАРТОН")

    plan = openpyxl.load_workbook(filename=plan, data_only=True)  # Opening source data with order, time, etc. info
    plan_s = plan["Plan"]

    for i in range(3, corrug_corrug.max_column):  # Определяем год
        val = corrug_corrug.cell(row=2, column=i).value
        if val is not None:
            if int(val) == int(current_year):
                break

    for j in range(i, corrug_corrug.max_column):  # Определяем месяц
        val = str(corrug_corrug.cell(row=1, column=j).value).strip()
        if val == "January":
            val = 1
        elif val == "February":
            val = 2
        elif val == "March":
            val = 3
        elif val == "April":
            val = 4
        elif val == "May":
            val = 5
        elif val == "June":
            val = 6
        elif val == "July":
            val = 7
        elif val == "August":
            val = 8
        elif val == "September":
            val = 9
        elif val == "October":
            val = 10
        elif val == "November":
            val = 11
        elif val == "December":
            val = 12
        if val is not None:
            if val in day_array[1]:
                break

    wk = corrug_corrug.cell(row=2, column=j).value
    for k in range(j, j + 45):  # Определяем день
        val = corrug_corrug.cell(row=3, column=k).value
        if corrug_corrug.cell(row=2, column=k).value is not None:
            wk = corrug_corrug.cell(row=2, column=k).value
        if current_year == current_year_real:
            if int(wk) < current_week or int(wk) > current_week + 2:
                continue
        if val in day_array[2]:
            for s in range(len(day_array[0])):
                if val == day_array[2][s]:
                    req_date[0].append(k)
                    req_date[1].append(day_array[2][s])
#####################################################################################################
    collected_data = [[], [], []]

    for n in range(len(day_array[3])):
        val = day_array[2][n]
        col = day_array[3][n]
        for o in range(4, plan_s.max_row):
            val_SKU = plan_s.cell(row=o, column=col).value
            neightbord_val_SKU = plan_s.cell(row=o, column=col + 2).value
            SKU_self = plan_s.cell(row=o, column=2).value

            if SKU_self is not None:
                if val_SKU is not None or neightbord_val_SKU is not None:
                    if val_SKU is None or val_SKU in ('', ' '):
                        val_SKU = 0
                    if neightbord_val_SKU is None or neightbord_val_SKU in ('', ' '):
                        neightbord_val_SKU = 0
                    collected_data[0].append(val)
                    collected_data[1].append(SKU_self)
                    collected_data[2].append(val_SKU + neightbord_val_SKU)
#####################################################################################################
    target_data = [[], [], []]

    for e in range(1, corrug_BOMmini.max_row):

        val = corrug_BOMmini.cell(row=e, column=2).value
        req_val = corrug_BOMmini.cell(row=e, column=1).value

        for a in range(len(collected_data[0])):

            temp_val = collected_data[1][a]

            if val == temp_val:
                target_data[0].append(collected_data[0][a])
                target_data[1].append(req_val)
                target_data[2].append(collected_data[2][a])
#####################################################################################################
    final_data = [[], [], []]

    for d in range(len(target_data[0])):

        act_data = target_data[0][d]
        act_sku = target_data[1][d]
        act_qnt = 0

        if d > 0:
            for t in range(len(final_data[0])):
                if act_data == final_data[0][t] and act_sku == final_data[1][t]:
                    act_data = '0'
                    act_sku = 0
                    break

        for f in range(len(target_data[0])):

            comparing_data = target_data[0][f]
            comparing_sku = target_data[1][f]

            if act_data == comparing_data and act_sku == comparing_sku:
                act_qnt += target_data[2][f]

        if act_qnt > 0:
            final_data[0].append(target_data[0][d])
            final_data[1].append(target_data[1][d])
            final_data[2].append(act_qnt)
#####################################################################################################
    for g in range(1, corrug_corrug.max_row):

        val = corrug_corrug.cell(row=g, column=2).value

        if str(val) in str(final_data[1]):

            for b in range(len(final_data[1])):

                act_sku = final_data[1][b]
                act_date = final_data[0][b]
                act_qnt = final_data[2][b]

                if str(val) == str(act_sku):
                    corrug_corrug.cell(row=g - 1,
                                    column=req_date[0][req_date[1].index(act_date)]).value = act_qnt
                    corrug_corrug.cell(row=g - 1,
                                    column=req_date[0][req_date[1].index(act_date)]).font = Font(
                        color='556B2F', size=9, name="Arial")

    corrug.save(corr)
    return True
